"""
每日新聞抓取腳本
來源：
  1. 臺北市政府首頁新聞稿 (gov.taipei)
  2. 台北服務通 (service.taipei) — 若有公告頁面
執行後在同目錄產出 news_YYYYMMDD.xlsx
"""

import re
import sys
from datetime import datetime, timedelta, timezone

import urllib3
import requests
from bs4 import BeautifulSoup

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── 設定 ─────────────────────────────────────────────────────────────────────
HOURS_BACK = 26          # 抓幾小時內的新聞（稍微多抓一點避免遺漏）
BASE_GOV = "https://www.gov.taipei"
NEWS_URL = (
    f"{BASE_GOV}/News.aspx"
    "?n=F0DDAF49B89E9413&sms=72544237BBE4C5F6&PageSize=100"
)

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0 Safari/537.36"
    ),
    "Accept-Language": "zh-TW,zh;q=0.9,en;q=0.8",
}
TW_TZ = timezone(timedelta(hours=8))
# ─────────────────────────────────────────────────────────────────────────────


def roc_to_date(roc_str: str) -> datetime | None:
    """將民國年日期字串轉為 datetime。接受 '115-06-03' 或 '115年06月03日'。"""
    roc_str = roc_str.strip()
    m = re.search(r"(\d{2,3})[-年](\d{1,2})[-月](\d{1,2})", roc_str)
    if not m:
        return None
    y, mo, d = int(m.group(1)) + 1911, int(m.group(2)), int(m.group(3))
    return datetime(y, mo, d, tzinfo=TW_TZ)


def fetch_gov_taipei(cutoff: datetime) -> list[dict]:
    """抓取 gov.taipei 新聞稿，回傳 cutoff 之後的項目。"""
    rows = []
    page = 1
    while True:
        url = NEWS_URL + f"&page={page}"
        try:
            resp = requests.get(url, headers=HEADERS, timeout=20, verify=False)
            resp.raise_for_status()
        except Exception as e:
            print(f"[gov.taipei] 第{page}頁抓取失敗：{e}", file=sys.stderr)
            break

        soup = BeautifulSoup(resp.text, "html.parser")
        table = soup.select_one("table.table") or soup.find("table")
        if not table:
            break

        trs = table.select("tbody tr") or table.select("tr")[1:]
        if not trs:
            break

        reached_old = False
        for tr in trs:
            tds = tr.find_all("td")
            if len(tds) < 3:
                continue

            # 找日期欄（通常第3欄）
            date_text = ""
            link_tag = None
            title_text = ""
            dept_text = ""

            for td in tds:
                text = td.get_text(strip=True)
                if re.search(r"\d{2,3}[-年]\d{1,2}[-月]\d{1,2}", text):
                    date_text = text
                a = td.find("a")
                if a and a.get_text(strip=True) and len(a.get_text(strip=True)) > 5:
                    link_tag = a
                    title_text = a.get_text(strip=True)

            # 發布機關：最後一個非空 td（排除日期欄和標題欄）
            non_empty = [td.get_text(strip=True) for td in tds if td.get_text(strip=True)]
            if non_empty:
                dept_text = non_empty[-1]
                if dept_text == title_text or re.search(r"\d{2,3}[-年]", dept_text):
                    dept_text = non_empty[-2] if len(non_empty) >= 2 else ""

            pub_date = roc_to_date(date_text)
            if pub_date is None:
                continue

            if pub_date < cutoff:
                reached_old = True
                continue

            href = ""
            if link_tag and link_tag.get("href"):
                raw_href = link_tag["href"]
                href = raw_href if raw_href.startswith("http") else f"{BASE_GOV}/{raw_href.lstrip('/')}"

            rows.append({
                "來源": "市府新聞稿",
                "日期": pub_date.strftime("%Y-%m-%d"),
                "標題": title_text,
                "發布機關": dept_text,
                "連結": href,
            })

        # 若整頁都是舊資料，停止翻頁
        if reached_old and not rows:
            break
        # 若這頁找到的都比 cutoff 舊，也停止
        if reached_old:
            break

        # 翻頁：若這頁筆數 < 100 表示最後一頁
        if len(trs) < 100:
            break
        page += 1

    return rows


def fetch_service_taipei(cutoff: datetime) -> list[dict]:
    """
    台北服務通目前是服務目錄型入口，沒有固定新聞清單頁。
    若未來有新聞 URL 可在此加入。
    """
    # TODO: 若 service.taipei 新增公告頁面，在此實作
    print("[service.taipei] 尚未找到新聞/公告列表頁面，此來源略過。", file=sys.stderr)
    return []


def write_excel(rows: list[dict], path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "新聞彙整"

    headers = ["來源", "日期", "標題", "發布機關", "連結"]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 交替列顏色
    fill_a = PatternFill("solid", fgColor="DCE6F1")
    fill_b = PatternFill("solid", fgColor="FFFFFF")

    for r, row in enumerate(rows, 2):
        fill = fill_a if r % 2 == 0 else fill_b
        for col, key in enumerate(headers, 1):
            cell = ws.cell(row=r, column=col, value=row.get(key, ""))
            cell.fill = fill
            cell.alignment = Alignment(vertical="center", wrap_text=(key == "標題"))
            if key == "連結" and row.get(key):
                cell.hyperlink = row[key]
                cell.font = Font(color="0563C1", underline="single")

    # 欄寬
    col_widths = [12, 12, 60, 25, 50]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(path)


def main():
    now = datetime.now(tz=TW_TZ)
    cutoff = now - timedelta(hours=HOURS_BACK)
    print(f"抓取範圍：{cutoff.strftime('%Y-%m-%d %H:%M')} ~ {now.strftime('%Y-%m-%d %H:%M')} (台灣時間)")

    all_rows: list[dict] = []

    print("正在抓取 市府新聞稿…")
    gov_rows = fetch_gov_taipei(cutoff)
    print(f"  → 找到 {len(gov_rows)} 筆")
    all_rows.extend(gov_rows)

    print("正在抓取 台北服務通…")
    svc_rows = fetch_service_taipei(cutoff)
    print(f"  → 找到 {len(svc_rows)} 筆")
    all_rows.extend(svc_rows)

    # 依日期降冪排序
    all_rows.sort(key=lambda x: x["日期"], reverse=True)

    if not all_rows:
        print("沒有符合時間範圍的新聞。")

    out_path = f"news_{now.strftime('%Y%m%d')}.xlsx"
    write_excel(all_rows, out_path)
    print(f"\n完成！共 {len(all_rows)} 筆，已存至：{out_path}")


if __name__ == "__main__":
    main()
