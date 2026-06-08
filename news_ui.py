"""
臺北市政府新聞稿抓取工具（UI 版）
功能：
  - 選擇起迄日期時間（預設：前一天 00:00 ~ 現在）
  - 抓取 gov.taipei 新聞稿列表 + 文章全文
  - 可選：透過 Ollama 產生摘要與 FAQ
  - 產出 Excel，檔名含執行時間戳記
"""

import re
import sys
import threading
import queue
import time
from datetime import datetime, timedelta, timezone
from pathlib import Path

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import tkcalendar

import urllib3
import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ── 常數 ──────────────────────────────────────────────────────────────────────
TW_TZ = timezone(timedelta(hours=8))
BASE_GOV = "https://www.gov.taipei"
NEWS_LIST_URL = (
    f"{BASE_GOV}/News.aspx"
    "?n=F0DDAF49B89E9413&sms=72544237BBE4C5F6&PageSize=100"
)
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"
    ),
    "Accept-Language": "zh-TW,zh;q=0.9",
}
OLLAMA_URL = "http://localhost:11434/api/generate"
# 打包成執行檔後 __file__ 會指向暫存目錄，改用 sys.executable 取得程式所在路徑
def _get_app_dir() -> Path:
    if getattr(sys, "frozen", False):          # Nuitka / PyInstaller 打包後
        return Path(sys.executable).parent
    return Path(__file__).parent               # 一般 Python 執行

OUT_DIR = _get_app_dir()
# ─────────────────────────────────────────────────────────────────────────────


def roc_to_date(roc_str: str) -> datetime | None:
    m = re.search(r"(\d{2,3})[-年](\d{1,2})[-月](\d{1,2})", roc_str.strip())
    if not m:
        return None
    y, mo, d = int(m.group(1)) + 1911, int(m.group(2)), int(m.group(3))
    try:
        return datetime(y, mo, d, tzinfo=TW_TZ)
    except ValueError:
        return None


def fetch_article_content(url: str) -> str:
    try:
        resp = requests.get(url, headers=HEADERS, timeout=20, verify=False)
        resp.encoding = "utf-8"
        soup = BeautifulSoup(resp.text, "html.parser")

        content_div = (
            soup.find("div", class_=re.compile(r"(news[-_]?content|article|content[-_]?body|rte)", re.I))
            or soup.find("article")
            or soup.find("div", id=re.compile(r"(content|article|main)", re.I))
        )
        if not content_div:
            h3 = soup.find("h3")
            content_div = h3.find_parent("div") if h3 else soup.find("body")

        if content_div:
            for tag in content_div.find_all(["script", "style", "nav", "header", "footer"]):
                tag.decompose()
            paras = content_div.find_all("p")
            text = "\n".join(p.get_text(" ", strip=True) for p in paras if p.get_text(strip=True))
            if not text:
                text = content_div.get_text("\n", strip=True)
            return text[:5000]
    except Exception as e:
        return f"[抓取失敗：{e}]"
    return ""


def ollama_available(model: str) -> bool:
    try:
        r = requests.get("http://localhost:11434/api/tags", timeout=3)
        tags = r.json().get("models", [])
        return any(model in t.get("name", "") for t in tags)
    except Exception:
        return False


def ollama_generate(prompt: str, model: str) -> str:
    try:
        resp = requests.post(
            OLLAMA_URL,
            json={"model": model, "prompt": prompt, "stream": False},
            timeout=300,
        )
        return resp.json().get("response", "").strip()
    except Exception as e:
        return f"[AI 失敗：{e}]"


def make_summary(content: str, model: str) -> str:
    if not content or content.startswith("["):
        return ""
    prompt = (
        "你是政府公文摘要助理，請用繁體中文將以下新聞稿濃縮成3～5句話的摘要，"
        "重點包含：事件主體、核心內容、時間地點（若有）。直接輸出摘要，不要加標題或說明。\n\n"
        f"{content[:3000]}"
    )
    return ollama_generate(prompt, model)


def make_faq(content: str, model: str) -> str:
    if not content or content.startswith("["):
        return ""
    prompt = (
        "你是政府新聞稿 FAQ 整理助理。請仔細閱讀以下新聞稿全文，"
        "用繁體中文依照下列固定格式逐項回答。"
        "每一項請直接從原文中找出對應資訊並摘錄，若原文未提及則寫「未提及」。\n\n"
        "Q1: 這則公告的主要內容是什麼？\n"
        "A1: （說明事件核心或政策重點，2～4句）\n\n"
        "Q2: 時間是什麼時候？\n"
        "A2: （施工期間、活動日期、截止日期、開放時間等，請完整摘錄原文中所有日期與時段）\n\n"
        "Q3: 地點在哪裡？\n"
        "A3: （請完整摘錄原文中所有路名、路口、巷弄、地址、區域，不可省略）\n\n"
        "Q4: 影響範圍或注意事項是什麼？\n"
        "A4: （封路範圍、替代路線、改道方式、管制措施等，逐點列出）\n\n"
        "Q5: 適用對象或申請資格是什麼？\n"
        "A5: （對象、身分條件、年齡限制、申請門檻等）\n\n"
        "Q6: 如何報名、申請或參與？\n"
        "A6: （報名方式、申請流程、所需文件、網址連結等）\n\n"
        "Q7: 費用是多少？有無補助或優惠？\n"
        "A7: （收費標準、免費項目、補助金額、優惠條件等）\n\n"
        "Q8: 如何聯絡或洽詢？\n"
        "A8: （電話、分機、網址、承辦單位，請完整列出）\n\n"
        "注意：\n"
        "- 地點（Q3）務必直接引用原文中的路名與地址，不可用「施工現場」或「附近」帶過。\n"
        "- 每個 A 的回答請盡量完整，不要過度精簡。\n"
        "直接輸出以上格式，不要加任何額外標題或說明。\n\n"
        "===新聞稿內容===\n"
        f"{content[:3000]}"
    )
    return ollama_generate(prompt, model)


def fetch_news_list(
    start_dt: datetime,
    end_dt: datetime,
    log_q: queue.Queue,
    use_ai: bool,
    ai_model: str,
    stop_flag: threading.Event | None = None,
) -> list[dict]:
    def stopped() -> bool:
        return stop_flag is not None and stop_flag.is_set()
    rows: list[dict] = []
    page = 1

    while True:
        if stopped():
            log_q.put("  [已停止]")
            return rows
        url = NEWS_LIST_URL + f"&page={page}"
        log_q.put(f"  抓取列表第 {page} 頁…")
        try:
            resp = requests.get(url, headers=HEADERS, timeout=20, verify=False)
            resp.encoding = "utf-8"
            resp.raise_for_status()
        except Exception as e:
            log_q.put(f"  [錯誤] 第{page}頁：{e}")
            break

        soup = BeautifulSoup(resp.text, "html.parser")
        table = soup.select_one("table.table") or soup.find("table")
        if not table:
            log_q.put("  找不到新聞列表表格，停止。")
            break

        trs = table.select("tbody tr") or table.select("tr")[1:]
        if not trs:
            break

        all_old = True

        for tr in trs:
            tds = tr.find_all("td")
            if len(tds) < 3:
                continue

            date_text = title_text = dept_text = ""
            link_tag = None

            for td in tds:
                text = td.get_text(strip=True)
                if re.search(r"\d{2,3}[-年]\d{1,2}[-月]\d{1,2}", text):
                    date_text = text
                a = td.find("a")
                if a and len(a.get_text(strip=True)) > 5:
                    link_tag = a
                    title_text = a.get_text(strip=True)

            non_empty = [td.get_text(strip=True) for td in tds if td.get_text(strip=True)]
            if non_empty:
                dept_text = non_empty[-1]
                if dept_text == title_text or re.search(r"\d{2,3}[-年]", dept_text):
                    dept_text = non_empty[-2] if len(non_empty) >= 2 else ""

            pub_date = roc_to_date(date_text)
            if pub_date is None:
                continue

            pub_day   = pub_date.replace(hour=0, minute=0, second=0)
            start_day = start_dt.replace(hour=0, minute=0, second=0)
            end_day   = end_dt.replace(hour=0, minute=0, second=0)

            if pub_day < start_day:
                continue
            if pub_day > end_day:
                continue

            all_old = False

            raw_href = link_tag["href"] if (link_tag and link_tag.get("href")) else ""
            full_href = (
                raw_href if raw_href.startswith("http")
                else f"{BASE_GOV}/{raw_href.lstrip('/')}"
            )

            rows.append({
                "來源":   "市府新聞稿",
                "日期":   pub_date.strftime("%Y-%m-%d"),
                "標題":   title_text,
                "發布機關": dept_text,
                "連結":   full_href,
                "內容":   "",
                "摘要":   "",
                "FAQ":    "",
            })

        if all_old and page > 1:
            break
        if len(trs) < 100:
            break
        page += 1

    # 逐篇抓正文，並視需要呼叫 AI
    total = len(rows)
    for i, row in enumerate(rows, 1):
        if stopped():
            log_q.put(f"  [已停止] 已完成 {i-1}/{total} 篇")
            return rows

        log_q.put(f"  [{i}/{total}] 抓內容：{row['標題'][:28]}…")
        if row["連結"]:
            row["內容"] = fetch_article_content(row["連結"])

        content = row["內容"]
        if not content:
            log_q.put(f"  [{i}/{total}] ⚠ 內容為空，跳過 AI")
        elif content.startswith("["):
            log_q.put(f"  [{i}/{total}] ⚠ 內容抓取失敗（{content[:40]}），跳過 AI")
        elif len(content) < 20:
            log_q.put(f"  [{i}/{total}] ⚠ 內容過短（{len(content)} 字），跳過 AI")
        time.sleep(0.3)

        can_ai = use_ai and content and not content.startswith("[") and len(content) >= 20
        if can_ai:
            if stopped():
                log_q.put(f"  [已停止] 已完成 {i-1}/{total} 篇")
                return rows
            log_q.put(f"  [{i}/{total}] AI 摘要中（內容 {len(content)} 字）…")
            result = make_summary(content, ai_model)
            if result.startswith("[AI 失敗"):
                log_q.put(f"  [{i}/{total}] ⚠ 摘要失敗：{result}")
            row["摘要"] = result

            if stopped():
                log_q.put(f"  [已停止] 已完成 {i-1}/{total} 篇")
                return rows
            log_q.put(f"  [{i}/{total}] AI FAQ 中…")
            result = make_faq(content, ai_model)
            if result.startswith("[AI 失敗"):
                log_q.put(f"  [{i}/{total}] ⚠ FAQ 失敗：{result}")
            row["FAQ"] = result

    return rows


def write_excel(rows: list[dict], path: Path,
                start_dt: datetime, end_dt: datetime) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "新聞彙整"

    # 欄位定義：(標題顯示名稱, row dict key 或 None=固定值, 固定值, 欄寬)
    # 新聞稿彙整日 = 執行當天；上架日/開始日期 = 新聞稿日期
    today_str = datetime.now(tz=TW_TZ).strftime("%Y-%m-%d")

    COLS = [
        ("新聞稿彙整日", None,       today_str,    14),
        ("新聞稿上架日", "日期",      None,         14),
        ("開始日期",     "日期",      None,         14),
        ("性質",         None,       "一般",        10),
        ("內容",         None,       "活動訊息",    14),
        ("主旨",         "標題",      None,         46),
        ("內容",         "內容",      None,         70),
        ("發布機關",     "發布機關",  None,         20),
        ("網址連結",     "連結",      None,         40),
        ("摘要",         "摘要",      None,         50),
        ("FAQ",          "FAQ",       None,         55),
    ]
    num_cols = len(COLS)

    # ── 第1列：標題說明列 ──
    title_text = (
        f"臺北市政府新聞稿彙整　"
        f"期間：{start_dt.strftime('%Y-%m-%d %H:%M')} ～ {end_dt.strftime('%Y-%m-%d %H:%M')}　"
        f"共 {len(rows)} 筆"
    )
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    title_cell = ws.cell(row=1, column=1, value=title_text)
    title_cell.font = Font(bold=True, size=12, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor="1F4E79")
    title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 24

    # ── 第2列：欄位標題 ──
    hdr_fill = PatternFill("solid", fgColor="2E75B6")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    for col, (hdr, _, __, ___) in enumerate(COLS, 1):
        c = ws.cell(row=2, column=col, value=hdr)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    # ── 第3列起：資料 ──
    fill_a = PatternFill("solid", fgColor="DCE6F1")
    fill_b = PatternFill("solid", fgColor="FFFFFF")

    for r, row in enumerate(rows, 3):
        fill = fill_a if r % 2 == 0 else fill_b
        for col, (hdr, key, default, _) in enumerate(COLS, 1):
            if key is None:
                val = default
            else:
                val = row.get(key, "")
            c = ws.cell(row=r, column=col, value=val)
            c.fill = fill
            c.alignment = Alignment(vertical="top", wrap_text=True)
            if hdr == "網址連結" and val:
                c.hyperlink = val
                c.font = Font(color="0563C1", underline="single")

    for i, (_, __, ___, width) in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(num_cols)}{ws.max_row}"

    wb.save(path)


# ── GUI ───────────────────────────────────────────────────────────────────────

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("臺北市政府新聞稿抓取工具")
        self.resizable(False, False)
        self._build_ui()
        self._set_defaults()
        self._check_ollama()

    def _build_ui(self):
        pad = {"padx": 10, "pady": 6}

        tk.Label(self, text="臺北市政府新聞稿抓取工具",
                 font=("微軟正黑體", 14, "bold"),
                 fg="#1F4E79").grid(row=0, column=0, columnspan=4, pady=(14, 4))

        # ── 快捷選項 ──
        quick_frame = tk.LabelFrame(self, text="快速選擇期間",
                                    font=("微軟正黑體", 9), padx=8, pady=4)
        quick_frame.grid(row=1, column=0, columnspan=4, sticky="we",
                         padx=10, pady=(0, 4))
        self._quick_var = tk.IntVar(value=1)
        opts = [
            (1, "1  當日（00:00 ~ 現在）"),
            (2, "2  前一日（全天）"),
            (3, "3  自訂區間"),
        ]
        for val, text in opts:
            ttk.Radiobutton(quick_frame, text=text, variable=self._quick_var,
                            value=val, command=self._on_quick).pack(
                            side="left", padx=10)

        # ── 起始 ──
        tk.Label(self, text="起始日期", font=("微軟正黑體", 10)).grid(
            row=2, column=0, sticky="e", **pad)
        self.cal_start = tkcalendar.DateEntry(self, width=12, date_pattern="yyyy/mm/dd",
                                              font=("微軟正黑體", 10))
        self.cal_start.grid(row=2, column=1, sticky="w", **pad)
        tk.Label(self, text="時間", font=("微軟正黑體", 10)).grid(
            row=2, column=2, sticky="e", padx=(0, 4))
        self.start_time = tk.StringVar(value="00:00")
        self._entry_start = ttk.Entry(self, textvariable=self.start_time, width=7,
                                      font=("微軟正黑體", 10))
        self._entry_start.grid(row=2, column=3, sticky="w", padx=(0, 10))

        # ── 結束 ──
        tk.Label(self, text="結束日期", font=("微軟正黑體", 10)).grid(
            row=3, column=0, sticky="e", **pad)
        self.cal_end = tkcalendar.DateEntry(self, width=12, date_pattern="yyyy/mm/dd",
                                            font=("微軟正黑體", 10))
        self.cal_end.grid(row=3, column=1, sticky="w", **pad)
        tk.Label(self, text="時間", font=("微軟正黑體", 10)).grid(
            row=3, column=2, sticky="e", padx=(0, 4))
        self.end_time = tk.StringVar()
        self._entry_end = ttk.Entry(self, textvariable=self.end_time, width=7,
                                    font=("微軟正黑體", 10))
        self._entry_end.grid(row=3, column=3, sticky="w", padx=(0, 10))

        # ── 輸出目錄 ──
        tk.Label(self, text="輸出目錄", font=("微軟正黑體", 10)).grid(
            row=4, column=0, sticky="e", **pad)
        self.out_dir = tk.StringVar(value=str(OUT_DIR))
        ttk.Entry(self, textvariable=self.out_dir, width=30,
                  font=("微軟正黑體", 9)).grid(row=4, column=1, columnspan=2,
                                               sticky="we", padx=(0, 4), pady=6)
        ttk.Button(self, text="瀏覽…", width=6,
                   command=self._browse_dir).grid(row=4, column=3,
                                                  sticky="w", padx=(0, 10))

        # AI 設定區
        sep = ttk.Separator(self, orient="horizontal")
        sep.grid(row=5, column=0, columnspan=4, sticky="we", padx=10, pady=(4, 0))

        self.use_ai = tk.BooleanVar(value=False)
        self.ai_chk = ttk.Checkbutton(self, text="啟用 AI（Ollama）產生摘要與 FAQ",
                                       variable=self.use_ai, command=self._toggle_ai)
        self.ai_chk.grid(row=6, column=0, columnspan=2, sticky="w", padx=12, pady=4)

        tk.Label(self, text="模型", font=("微軟正黑體", 10)).grid(
            row=6, column=2, sticky="e", padx=(0, 4))
        self.ai_model = tk.StringVar(value="llama3.2:3b")
        self.model_combo = ttk.Combobox(self, textvariable=self.ai_model, width=16,
                                        font=("微軟正黑體", 10), state="disabled")
        self.model_combo.grid(row=6, column=3, sticky="w", padx=(0, 10))

        self.ollama_status = tk.Label(self, text="● Ollama 檢查中…",
                                      font=("微軟正黑體", 9), fg="gray")
        self.ollama_status.grid(row=7, column=0, columnspan=4, sticky="w", padx=14, pady=(0, 4))

        sep2 = ttk.Separator(self, orient="horizontal")
        sep2.grid(row=8, column=0, columnspan=4, sticky="we", padx=10, pady=(0, 4))

        # 開始 / 停止按鈕
        btn_frame = tk.Frame(self)
        btn_frame.grid(row=9, column=0, columnspan=4, pady=6)
        self.btn = ttk.Button(btn_frame, text="▶ 開始抓取", command=self._start)
        self.btn.pack(side="left", ipadx=20, ipady=4, padx=6)
        self.btn_stop = ttk.Button(btn_frame, text="■ 停止", command=self._stop,
                                   state="disabled")
        self.btn_stop.pack(side="left", ipadx=14, ipady=4, padx=6)

        # 進度條
        self.progress = ttk.Progressbar(self, mode="indeterminate", length=480)
        self.progress.grid(row=10, column=0, columnspan=4, padx=10, pady=(0, 4))

        # 日誌
        tk.Label(self, text="執行紀錄", font=("微軟正黑體", 10, "bold")).grid(
            row=11, column=0, columnspan=4, sticky="w", padx=12)
        frame = tk.Frame(self)
        frame.grid(row=12, column=0, columnspan=4, padx=10, pady=(0, 12), sticky="nsew")
        self.log_box = tk.Text(frame, height=14, width=64, state="disabled",
                               font=("Consolas", 9), bg="#F5F5F5", relief="sunken")
        sb = ttk.Scrollbar(frame, command=self.log_box.yview)
        self.log_box.config(yscrollcommand=sb.set)
        self.log_box.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")

    def _browse_dir(self):
        d = filedialog.askdirectory(initialdir=self.out_dir.get(), title="選擇輸出目錄")
        if d:
            self.out_dir.set(d)

    def _set_defaults(self):
        self._quick_var.set(1)
        self._on_quick()

    def _on_quick(self):
        """依快捷選項設定日期時間，並控制自訂欄位的啟用狀態。"""
        now = datetime.now(tz=TW_TZ)
        choice = self._quick_var.get()

        if choice == 1:          # 當日 00:00 ~ 現在
            self.cal_start.config(state="normal")
            self.cal_end.config(state="normal")
            self.cal_start.set_date(now.date())
            self.start_time.set("00:00")
            self.cal_end.set_date(now.date())
            self.end_time.set(now.strftime("%H:%M"))
            self._set_custom_state("disabled")

        elif choice == 2:        # 前一日全天
            yesterday = now - timedelta(days=1)
            self.cal_start.config(state="normal")
            self.cal_end.config(state="normal")
            self.cal_start.set_date(yesterday.date())
            self.start_time.set("00:00")
            self.cal_end.set_date(yesterday.date())
            self.end_time.set("23:59")
            self._set_custom_state("disabled")

        else:                    # 自訂
            self._set_custom_state("normal")

    def _set_custom_state(self, state: str):
        """控制日期時間欄位是否可編輯。"""
        cal_state = "normal" if state == "normal" else "disabled"
        self.cal_start.config(state=cal_state)
        self.cal_end.config(state=cal_state)
        self._entry_start.config(state=state)
        self._entry_end.config(state=state)

    def _get_ollama_models(self) -> list[str]:
        """從 Ollama API 取得已安裝的模型清單。"""
        try:
            r = requests.get("http://localhost:11434/api/tags", timeout=3)
            return [m["name"] for m in r.json().get("models", [])]
        except Exception:
            return []

    def _check_ollama(self):
        """背景確認 Ollama 是否在線，並更新模型下拉清單。"""
        def check():
            models = self._get_ollama_models()
            if models:
                # 更新下拉清單
                self.model_combo["values"] = models
                # 若目前選的不在清單裡，自動選第一個
                if self.ai_model.get() not in models:
                    self.ai_model.set(models[0])
                self.ollama_status.config(
                    text=f"● Ollama 已連線，共 {len(models)} 個模型可用", fg="#217346")
            else:
                self.model_combo["values"] = []
                try:
                    requests.get("http://localhost:11434", timeout=2)
                    self.ollama_status.config(
                        text="● Ollama 已啟動，但尚未下載任何模型（請執行 ollama pull llama3.2:3b）",
                        fg="#C55A11")
                except Exception:
                    self.ollama_status.config(
                        text="● Ollama 未啟動（請安裝並執行 ollama serve）", fg="#C00000")
        threading.Thread(target=check, daemon=True).start()

    def _toggle_ai(self):
        state = "readonly" if self.use_ai.get() else "disabled"
        self.model_combo.config(state=state)
        if self.use_ai.get():
            self._check_ollama()

    def _log(self, msg: str):
        self.log_box.config(state="normal")
        self.log_box.insert("end", msg + "\n")
        self.log_box.see("end")
        self.log_box.config(state="disabled")

    def _parse_dt(self, date_entry, time_str: str) -> datetime | None:
        try:
            d = date_entry.get_date()
            h, m = map(int, time_str.strip().split(":"))
            return datetime(d.year, d.month, d.day, h, m, tzinfo=TW_TZ)
        except Exception:
            return None

    def _start(self):
        start_dt = self._parse_dt(self.cal_start, self.start_time.get())
        end_dt   = self._parse_dt(self.cal_end,   self.end_time.get())

        if not start_dt or not end_dt:
            messagebox.showerror("格式錯誤", "時間格式請輸入 HH:MM，例如 09:30")
            return
        if start_dt >= end_dt:
            messagebox.showerror("時間錯誤", "起始時間必須早於結束時間")
            return

        use_ai   = self.use_ai.get()
        ai_model = self.ai_model.get().strip()

        if use_ai and not ollama_available(ai_model):
            if not messagebox.askyesno("Ollama 無法連線",
                                       f"找不到模型 {ai_model}，\n要繼續（不產生 AI 欄位）嗎？"):
                return
            use_ai = False

        self._stop_flag = threading.Event()
        self.btn.config(state="disabled")
        self.btn_stop.config(state="normal")
        self.progress.start(10)
        self._log(f"開始抓取：{start_dt.strftime('%Y-%m-%d %H:%M')} ~ {end_dt.strftime('%Y-%m-%d %H:%M')}")
        if use_ai:
            self._log(f"  AI 功能開啟，模型：{ai_model}")

        log_q: queue.Queue = queue.Queue()
        result: dict = {}

        def worker():
            try:
                rows = fetch_news_list(start_dt, end_dt, log_q, use_ai, ai_model,
                                       stop_flag=self._stop_flag)
                result["rows"] = rows
                result["error"] = None
            except Exception as e:
                result["rows"] = []
                result["error"] = str(e)
            log_q.put(None)

        threading.Thread(target=worker, daemon=True).start()
        self._poll(log_q, result, start_dt, end_dt)

    def _stop(self):
        if hasattr(self, "_stop_flag"):
            self._stop_flag.set()
        self.btn_stop.config(state="disabled")
        self._log("正在停止，等待目前工作完成…")

    def _poll(self, log_q: queue.Queue, result: dict, start_dt, end_dt):
        try:
            while True:
                msg = log_q.get_nowait()
                if msg is None:
                    self.progress.stop()
                    self.btn.config(state="normal")
                    self.btn_stop.config(state="disabled")
                    if result.get("error"):
                        self._log(f"[錯誤] {result['error']}")
                        messagebox.showerror("錯誤", result["error"])
                    else:
                        rows = result["rows"]
                        self._log(f"共找到 {len(rows)} 筆新聞")
                        if rows:
                            self._save(rows, start_dt, end_dt)
                    return
                self._log(msg)
        except queue.Empty:
            pass
        self.after(100, lambda: self._poll(log_q, result, start_dt, end_dt))

    def _save(self, rows: list[dict], start_dt: datetime, end_dt: datetime):
        now = datetime.now(tz=TW_TZ)
        # 檔名：news_起始日期_結束日期_產生時間.xlsx
        s = start_dt.strftime("%Y%m%d")
        e = end_dt.strftime("%Y%m%d")
        t = now.strftime("%Y%m%d_%H%M")
        fname = f"news_{s}_{e}_{t}.xlsx" if s != e else f"news_{s}_{t}.xlsx"
        out_path = Path(self.out_dir.get()) / fname
        try:
            write_excel(rows, out_path, start_dt, end_dt)
            self._log(f"✓ 已儲存：{out_path}")
            messagebox.showinfo("完成", f"共 {len(rows)} 筆新聞\n已儲存至：\n{out_path}")
        except Exception as e:
            self._log(f"[儲存失敗] {e}")
            messagebox.showerror("儲存失敗", str(e))


if __name__ == "__main__":
    app = App()
    app.mainloop()
